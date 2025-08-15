# Monthly Reporting System
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Sum, Q, Avg
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
import pandas as pd
from io import BytesIO
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from sessions.models import TherapySession
from therapists.models import TherapistProfile
from clients.models import ClientProfile
from companies.models import Company, CompanyReport
from payments.models import Payment
from coupons.models import IndividualCoupon


class MonthlyReportGenerator:
    """
    Generate comprehensive monthly reports for admin portal
    """
    
    def __init__(self, year, month):
        self.year = year
        self.month = month
        self.start_date = datetime(year, month, 1).date()
        if month == 12:
            self.end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            self.end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    def generate_overall_report(self):
        """
        Generate overall platform monthly report
        """
        report_data = {
            'period': f"{self.start_date.strftime('%B %Y')}",
            'start_date': self.start_date,
            'end_date': self.end_date,
            'clients_data': self._get_clients_data(),
            'therapists_data': self._get_therapists_data(),
            'sessions_data': self._get_sessions_data(),
            'financial_data': self._get_financial_data(),
            'companies_data': self._get_companies_data(),
            'coupons_data': self._get_coupons_data(),
        }
        
        return report_data
    
    def _get_clients_data(self):
        """Get client-related statistics"""
        # Total clients seen
        clients_seen = ClientProfile.objects.filter(
            sessions__scheduled_date__range=[self.start_date, self.end_date],
            sessions__status='completed'
        ).distinct().count()
        
        # New clients registered
        new_clients = ClientProfile.objects.filter(
            created_at__date__range=[self.start_date, self.end_date]
        ).count()
        
        # Active clients (had at least one session)
        active_clients = ClientProfile.objects.filter(
            sessions__scheduled_date__range=[self.start_date, self.end_date]
        ).distinct().count()
        
        return {
            'clients_seen': clients_seen,
            'new_clients': new_clients,
            'active_clients': active_clients,
        }
    
    def _get_therapists_data(self):
        """Get therapist-related statistics"""
        # Active therapists (conducted at least one session)
        active_therapists = TherapistProfile.objects.filter(
            sessions__scheduled_date__range=[self.start_date, self.end_date],
            sessions__status='completed'
        ).distinct()
        
        therapist_stats = []
        for therapist in active_therapists:
            sessions_count = therapist.sessions.filter(
                scheduled_date__range=[self.start_date, self.end_date],
                status='completed'
            ).count()
            
            clients_count = therapist.sessions.filter(
                scheduled_date__range=[self.start_date, self.end_date],
                status='completed'
            ).values('client').distinct().count()
            
            revenue = therapist.payments.filter(
                payment_date__date__range=[self.start_date, self.end_date],
                status='completed'
            ).aggregate(total=Sum('final_amount'))['total'] or 0
            
            therapist_stats.append({
                'therapist_name': therapist.user.get_full_name(),
                'therapist_id': therapist.id,
                'sessions_conducted': sessions_count,
                'clients_served': clients_count,
                'revenue_generated': revenue,
            })
        
        return {
            'active_therapists_count': active_therapists.count(),
            'therapist_details': therapist_stats,
        }
    
    def _get_sessions_data(self):
        """Get session-related statistics"""
        sessions = TherapySession.objects.filter(
            scheduled_date__range=[self.start_date, self.end_date]
        )
        
        total_sessions = sessions.count()
        completed_sessions = sessions.filter(status='completed').count()
        cancelled_sessions = sessions.filter(status='cancelled').count()
        no_show_sessions = sessions.filter(status='no_show').count()
        
        # Session types breakdown
        session_types = sessions.values('session_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return {
            'total_sessions': total_sessions,
            'completed_sessions': completed_sessions,
            'cancelled_sessions': cancelled_sessions,
            'no_show_sessions': no_show_sessions,
            'completion_rate': (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0,
            'session_types': list(session_types),
        }
    
    def _get_financial_data(self):
        """Get financial statistics"""
        payments = Payment.objects.filter(
            payment_date__date__range=[self.start_date, self.end_date],
            status='completed'
        )
        
        total_revenue = payments.aggregate(total=Sum('final_amount'))['total'] or 0
        total_discount = payments.aggregate(total=Sum('discount_amount'))['total'] or 0
        
        # Revenue by payment type
        revenue_by_type = payments.values('payment_type').annotate(
            total=Sum('final_amount'),
            count=Count('id')
        ).order_by('-total')
        
        # Company vs direct client revenue
        company_revenue = payments.filter(company__isnull=False).aggregate(
            total=Sum('final_amount')
        )['total'] or 0
        
        direct_revenue = payments.filter(company__isnull=True).aggregate(
            total=Sum('final_amount')
        )['total'] or 0
        
        return {
            'total_revenue': total_revenue,
            'total_discount': total_discount,
            'company_revenue': company_revenue,
            'direct_revenue': direct_revenue,
            'revenue_by_type': list(revenue_by_type),
        }
    
    def _get_companies_data(self):
        """Get company-related statistics"""
        companies_with_sessions = Company.objects.filter(
            payments__payment_date__date__range=[self.start_date, self.end_date],
            payments__status='completed'
        ).distinct()
        
        company_stats = []
        for company in companies_with_sessions:
            sessions_count = TherapySession.objects.filter(
                payment__company=company,
                scheduled_date__range=[self.start_date, self.end_date],
                status='completed'
            ).count()
            
            revenue = company.payments.filter(
                payment_date__date__range=[self.start_date, self.end_date],
                status='completed'
            ).aggregate(total=Sum('final_amount'))['total'] or 0
            
            discount = company.payments.filter(
                payment_date__date__range=[self.start_date, self.end_date],
                status='completed'
            ).aggregate(total=Sum('discount_amount'))['total'] or 0
            
            employees_served = ClientProfile.objects.filter(
                company=company,
                sessions__scheduled_date__range=[self.start_date, self.end_date],
                sessions__status='completed'
            ).distinct().count()
            
            company_stats.append({
                'company_name': company.name,
                'company_id': company.id,
                'sessions_count': sessions_count,
                'employees_served': employees_served,
                'revenue': revenue,
                'discount_given': discount,
            })
        
        return {
            'active_companies_count': companies_with_sessions.count(),
            'company_details': company_stats,
        }
    
    def _get_coupons_data(self):
        """Get coupon usage statistics"""
        coupons_used = IndividualCoupon.objects.filter(
            used_at__date__range=[self.start_date, self.end_date],
            status='used'
        )
        
        total_coupons_used = coupons_used.count()
        total_discount_given = 0
        
        # Calculate total discount from coupon usage
        for coupon in coupons_used:
            if coupon.coupon_system.discount_type == 'percentage':
                # This would need to be calculated based on actual session cost
                pass
            elif coupon.coupon_system.discount_type == 'fixed_amount':
                total_discount_given += coupon.coupon_system.discount_value
        
        return {
            'total_coupons_used': total_coupons_used,
            'total_discount_given': total_discount_given,
        }
    
    def generate_excel_report(self, report_data):
        """Generate Excel report"""
        output = BytesIO()
        workbook = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        
        # Add summary data
        summary_sheet['A1'] = f"Monthly Report - {report_data['period']}"
        summary_sheet['A3'] = "Clients Seen"
        summary_sheet['B3'] = report_data['clients_data']['clients_seen']
        summary_sheet['A4'] = "Total Sessions"
        summary_sheet['B4'] = report_data['sessions_data']['total_sessions']
        summary_sheet['A5'] = "Total Revenue"
        summary_sheet['B5'] = report_data['financial_data']['total_revenue']
        summary_sheet['A6'] = "Active Therapists"
        summary_sheet['B6'] = report_data['therapists_data']['active_therapists_count']
        
        # Therapist details sheet
        therapist_sheet = workbook.create_sheet("Therapists")
        therapist_sheet['A1'] = "Therapist Name"
        therapist_sheet['B1'] = "Sessions Conducted"
        therapist_sheet['C1'] = "Clients Served"
        therapist_sheet['D1'] = "Revenue Generated"
        
        for idx, therapist in enumerate(report_data['therapists_data']['therapist_details'], 2):
            therapist_sheet[f'A{idx}'] = therapist['therapist_name']
            therapist_sheet[f'B{idx}'] = therapist['sessions_conducted']
            therapist_sheet[f'C{idx}'] = therapist['clients_served']
            therapist_sheet[f'D{idx}'] = therapist['revenue_generated']
        
        # Company details sheet
        if report_data['companies_data']['company_details']:
            company_sheet = workbook.create_sheet("Companies")
            company_sheet['A1'] = "Company Name"
            company_sheet['B1'] = "Sessions"
            company_sheet['C1'] = "Employees Served"
            company_sheet['D1'] = "Revenue"
            company_sheet['E1'] = "Discount Given"
            
            for idx, company in enumerate(report_data['companies_data']['company_details'], 2):
                company_sheet[f'A{idx}'] = company['company_name']
                company_sheet[f'B{idx}'] = company['sessions_count']
                company_sheet[f'C{idx}'] = company['employees_served']
                company_sheet[f'D{idx}'] = company['revenue']
                company_sheet[f'E{idx}'] = company['discount_given']
        
        workbook.save(output)
        output.seek(0)
        return output
    
    def generate_pdf_report(self, report_data):
        """Generate PDF report"""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(f"Monthly Report - {report_data['period']}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Summary section
        summary_data = [
            ['Metric', 'Value'],
            ['Clients Seen', str(report_data['clients_data']['clients_seen'])],
            ['Total Sessions', str(report_data['sessions_data']['total_sessions'])],
            ['Completed Sessions', str(report_data['sessions_data']['completed_sessions'])],
            ['Total Revenue', f"₹{report_data['financial_data']['total_revenue']:,.2f}"],
            ['Active Therapists', str(report_data['therapists_data']['active_therapists_count'])],
            ['Active Companies', str(report_data['companies_data']['active_companies_count'])],
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Therapist performance section
        if report_data['therapists_data']['therapist_details']:
            therapist_title = Paragraph("Therapist Performance", styles['Heading2'])
            story.append(therapist_title)
            story.append(Spacer(1, 10))
            
            therapist_data = [['Therapist', 'Sessions', 'Clients', 'Revenue']]
            for therapist in report_data['therapists_data']['therapist_details'][:10]:  # Top 10
                therapist_data.append([
                    therapist['therapist_name'],
                    str(therapist['sessions_conducted']),
                    str(therapist['clients_served']),
                    f"₹{therapist['revenue_generated']:,.2f}"
                ])
            
            therapist_table = Table(therapist_data)
            therapist_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(therapist_table)
        
        doc.build(story)
        output.seek(0)
        return output
    
    def save_company_reports(self, report_data):
        """Save individual company reports"""
        for company_data in report_data['companies_data']['company_details']:
            company = Company.objects.get(id=company_data['company_id'])
            
            # Create or update company report
            report, created = CompanyReport.objects.get_or_create(
                company=company,
                report_type='monthly',
                period_start=self.start_date,
                period_end=self.end_date,
                defaults={
                    'total_employees_registered': company_data['employees_served'],
                    'total_sessions_conducted': company_data['sessions_count'],
                    'total_amount_spent': company_data['revenue'],
                    'total_discount_given': company_data['discount_given'],
                    'is_generated': True,
                    'generated_at': timezone.now(),
                }
            )
            
            if not created:
                # Update existing report
                report.total_employees_registered = company_data['employees_served']
                report.total_sessions_conducted = company_data['sessions_count']
                report.total_amount_spent = company_data['revenue']
                report.total_discount_given = company_data['discount_given']
                report.is_generated = True
                report.generated_at = timezone.now()
                report.save()